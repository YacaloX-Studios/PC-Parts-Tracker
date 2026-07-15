"""
PC Parts Tracker - Conversor Excel a JSON

Lee el Excel del usuario y genera un archivo JSON con todos los productos.
Se puede re-ejecutar para actualizar los precios desde el Excel.
"""

import json
import re
import sys
from pathlib import Path
from datetime import datetime

sys.path.insert(0, str(Path(__file__).parent))

import openpyxl
from utils.excel import infer_category, extract_url_from_hyperlink


def convert_excel_to_json(
    excel_path: str | None = None,
    output_path: str | None = None,
) -> list[dict]:
    """
    Convierte el Excel a un archivo JSON con precios.

    Args:
        excel_path: Ruta al archivo Excel. Si es None, busca automaticamente.
        output_path: Ruta de salida JSON. Si es None, usa data/prices.json.

    Formato de salida:
    [
        {
            "name": "ASUS Dual RTX 5060",
            "category": "GPU",
            "brand": "ASUS",
            "store": "MercadoLibre",
            "url": "https://...",
            "price": 1069956,
            "currency": "COP",
            "last_updated": "2026-07-14"
        },
        ...
    ]
    """
    project_root = Path(__file__).parent
    if excel_path is None:
        candidates = list(project_root.glob("*.xlsm")) + list(project_root.glob("*.xlsx"))
        if not candidates:
            print("No se encontro archivo Excel en el directorio del proyecto")
            return []
        excel_path = str(candidates[0])
    if output_path is None:
        output_path = str(project_root / "data" / "prices.json")

    print(f"Leyendo: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, read_only=False, keep_vba=True)
    ws = wb.active

    products = []

    for row_num in range(4, 30):
        name_cell = ws[f"B{row_num}"]
        price_cell = ws[f"C{row_num}"]
        url_cell = ws[f"E{row_num}"]
        store_cell = ws[f"F{row_num}"]

        name = name_cell.value
        if not name or str(name).strip().upper() == "TOTAL":
            continue

        name = str(name).strip()

        price = None
        if price_cell and price_cell.value is not None:
            if isinstance(price_cell.value, (int, float)):
                price = float(price_cell.value)
            elif isinstance(price_cell.value, str):
                cleaned = re.sub(r"[^\d.,]", "", price_cell.value)
                if cleaned:
                    try:
                        price = float(cleaned.replace(",", ""))
                    except ValueError:
                        pass

        url = extract_url_from_hyperlink(url_cell) if url_cell else None

        store = "MercadoLibre"
        if store_cell and store_cell.value:
            if isinstance(store_cell.value, str) and store_cell.value.strip():
                store = store_cell.value.strip()

        category = infer_category(name)
        parts = name.split()
        brand = parts[0] if parts else None
        model = " ".join(parts[1:]) if len(parts) > 1 else None

        products.append({
            "name": name,
            "category": category,
            "brand": brand,
            "model": model,
            "store": store,
            "url": url or "",
            "price": price,
            "currency": "COP",
            "last_updated": datetime.now().strftime("%Y-%m-%d"),
        })

        status = f"${price:,.0f}" if price else "Sin precio"
        print(f"  {name:<45} {category:<15} {status}")

    wb.close()

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    with open(output, "w", encoding="utf-8") as f:
        json.dump(products, f, indent=2, ensure_ascii=False)

    print(f"\nGuardado: {output_path}")
    print(f"Total: {len(products)} productos")
    return products


if __name__ == "__main__":
    convert_excel_to_json()
