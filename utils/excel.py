"""
PC Parts Tracker - Gestión de Excel/CSV

Importa y exporta datos de productos usando openpyxl y pandas.
Incluye migración automática desde el Excel existente del usuario.
"""

import re
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd

from database.models import Product
from utils.logger import get_logger

logger = get_logger("utils.excel")

CATEGORY_RULES: list[tuple[list[str], str]] = [
    (["rtx", "gtx", "rx", "intel arc", "geforce", "radeon", "tarjeta grafica"], "GPU"),
    (["ryzen", "intel core", "i5", "i7", "i9", "athlon", "phenom", "cpu"], "CPU"),
    (["ddr4", "ddr5", "ram", "memoria"], "RAM"),
    (["b650", "b550", "x670", "x570", "b460", "b660", "z690", "z790", "a620", "a520", "placa", "motherboard"], "Motherboard"),
    (["ssd", "nvme", "p3 plus", "evo", "970", "980", "sn770", "kingston"], "SSD"),
    (["hdd", "wd blue", "seagate", "barracuda"], "HDD"),
    (["psu", "fuente", "650w", "750w", "550w", "850w", "alimentacion", "rm850", "rm750", "rm650", "rm1000", "evga", "seasonic", "cx650", "cx750"], "PSU"),
    (["gabinete", "case", "air 100", "mesh", "corsair 4000", "nzxt"], "Case"),
    (["cooler", "disipador", "assassin", "thermalright", "noctua", "deepcool", "water cooler"], "Cooler"),
    (["monitor", "24\"", "27\"", "144hz", "165hz", "ips"], "Monitor"),
    (["teclado", "keyboard", "mouse", "mousepad", "webcam", "audifono", "parlantes", "usb"], "Perifericos"),
    (["wifi", "bluetooth", "fenvi", "pcie", "tarjeta red"], "Perifericos"),
]


def infer_category(product_name: str) -> str:
    """
    Infiera la categoría de un producto basándose en su nombre.

    Args:
        product_name: Nombre del producto.

    Returns:
        Categoría detectada o "Sin categoria".
    """
    name_lower = product_name.lower()
    for keywords, category in CATEGORY_RULES:
        for keyword in keywords:
            if keyword in name_lower:
                return category
    return "Sin categoria"


def extract_url_from_hyperlink(cell) -> str | None:
    """
    Extrae una URL de una celda de Excel que puede ser HYPERLINK.

    Maneja:
      - =HYPERLINK("url","texto")
      - Celdas con valor URL directa
    """
    if cell.value is None:
        return None

    if isinstance(cell.value, str):
        match = re.search(r'HYPERLINK\s*\(\s*"([^"]+)"', cell.value)
        if match:
            return match.group(1)
        if cell.value.startswith("http"):
            return cell.value

    if cell.hyperlink:
        return cell.hyperlink.target

    return None


class ExcelManager:
    """Gestor de importación/exportación de Excel y CSV."""

    def migrate_from_existing_excel(self, filepath: str) -> list[dict]:
        """
        Migra productos desde el Excel existente "Cotización PC Gamer.xlsm".

        Formato esperado:
          Columna B: Componente (nombre)
          Columna C: Precio
          Columna E: URL (HYPERLINK)
          Columna F: Tienda

        Args:
            filepath: Ruta al archivo Excel.

        Returns:
            Lista de dicts listos para crear Product.
        """
        logger.info("Iniciando migración desde: %s", filepath)
        products_data: list[dict] = []

        try:
            wb = openpyxl.load_workbook(filepath, read_only=False, keep_vba=True)
            ws = wb.active

            for row in ws.iter_rows(min_row=4, max_row=20, values_only=False):
                name_cell = row[1]  # Columna B
                price_cell = row[2]  # Columna C
                url_cell = row[4] if len(row) > 4 else None  # Columna E
                store_cell = row[5] if len(row) > 5 else None  # Columna F

                name = name_cell.value
                if not name or str(name).strip().upper() == "TOTAL":
                    continue

                name = str(name).strip()

                price = None
                if price_cell and price_cell.value:
                    if isinstance(price_cell.value, (int, float)):
                        price = float(price_cell.value)
                    elif isinstance(price_cell.value, str):
                        cleaned = re.sub(r"[^\d.,]", "", price_cell.value)
                        if cleaned:
                            try:
                                price = float(cleaned.replace(",", ""))
                            except ValueError:
                                pass

                url = extract_url_from_hyperlink(url_cell) if url_cell else None
                if not url:
                    continue

                store = "MercadoLibre"
                if store_cell and store_cell.value:
                    if isinstance(store_cell.value, str) and store_cell.value.strip():
                        store = store_cell.value.strip()

                category = infer_category(name)

                brand = None
                model = None
                parts = name.split()
                if parts:
                    brand = parts[0]
                    model = " ".join(parts[1:]) if len(parts) > 1 else None

                products_data.append({
                    "name": name,
                    "category": category,
                    "brand": brand,
                    "model": model,
                    "store": store,
                    "url": url,
                    "current_price": price,
                    "previous_price": None,
                    "lowest_price": price,
                    "highest_price": price,
                    "currency": "COP",
                    "target_price": None,
                    "active": True,
                })

            wb.close()
            logger.info(
                "Migración completada: %d productos extraídos", len(products_data)
            )

        except Exception as e:
            logger.error("Error durante la migración: %s", e)
            raise

        return products_data

    def import_from_excel(self, filepath: str) -> list[dict]:
        """
        Importa productos desde un Excel genérico.

        Formato esperado:
          Nombre | Categoría | Marca | Tienda | URL | Precio

        Args:
            filepath: Ruta al archivo Excel.

        Returns:
            Lista de dicts listos para crear Product.
        """
        logger.info("Importando desde Excel: %s", filepath)
        products_data: list[dict] = []

        try:
            wb = openpyxl.load_workbook(filepath, read_only=True)
            ws = wb.active

            headers = []
            for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=False), []):
                headers.append(str(cell.value).lower().strip() if cell.value else "")

            for row in ws.iter_rows(min_row=2, max_row=50, values_only=True):
                row_dict = dict(zip(headers, row))

                name = row_dict.get("nombre") or row_dict.get("name") or row_dict.get("producto")
                if not name:
                    continue

                price = None
                for key in ["precio", "price", "precio actual"]:
                    if key in row_dict and row_dict[key]:
                        try:
                            price = float(str(row_dict[key]).replace(",", "").replace("$", "").strip())
                        except (ValueError, TypeError):
                            pass

                url = row_dict.get("url") or row_dict.get("enlace") or row_dict.get("link") or ""

                store = row_dict.get("tienda") or row_dict.get("store") or "Desconocida"
                category = row_dict.get("categoria") or row_dict.get("category") or infer_category(str(name))

                products_data.append({
                    "name": str(name).strip(),
                    "category": str(category).strip(),
                    "brand": row_dict.get("marca") or row_dict.get("brand"),
                    "model": row_dict.get("modelo") or row_dict.get("model"),
                    "store": str(store).strip(),
                    "url": str(url).strip(),
                    "current_price": price,
                    "previous_price": None,
                    "lowest_price": price,
                    "highest_price": price,
                    "currency": "COP",
                    "target_price": None,
                    "active": True,
                })

            wb.close()
            logger.info("Importación completada: %d productos", len(products_data))

        except Exception as e:
            logger.error("Error importando Excel: %s", e)
            raise

        return products_data

    def export_to_excel(self, products: list[Product], filepath: str) -> None:
        """
        Exporta productos a un archivo Excel.

        Args:
            products: Lista de productos a exportar.
            filepath: Ruta del archivo de salida.
        """
        logger.info("Exportando %d productos a Excel: %s", len(products), filepath)

        data = []
        for p in products:
            data.append({
                "ID": p.id,
                "Nombre": p.name,
                "Categoria": p.category,
                "Marca": p.brand or "",
                "Modelo": p.model or "",
                "Tienda": p.store,
                "URL": p.url,
                "Precio Actual": p.current_price or 0,
                "Precio Anterior": p.previous_price or 0,
                "Minimo Historico": p.lowest_price or 0,
                "Maximo Historico": p.highest_price or 0,
                "Moneda": p.currency,
                "Precio Objetivo": p.target_price or "",
                "Ultima Actualizacion": p.last_updated.strftime("%Y-%m-%d %H:%M") if p.last_updated else "",
                "Fecha Creacion": p.created_at.strftime("%Y-%m-%d") if p.created_at else "",
                "Activo": "Si" if p.active else "No",
            })

        df = pd.DataFrame(data)
        df.to_excel(filepath, index=False, engine="openpyxl")
        logger.info("Exportación Excel completada: %s", filepath)

    def export_to_csv(self, products: list[Product], filepath: str) -> None:
        """
        Exporta productos a un archivo CSV.

        Args:
            products: Lista de productos a exportar.
            filepath: Ruta del archivo de salida.
        """
        logger.info("Exportando %d productos a CSV: %s", len(products), filepath)

        data = []
        for p in products:
            data.append({
                "ID": p.id,
                "Nombre": p.name,
                "Categoria": p.category,
                "Marca": p.brand or "",
                "Modelo": p.model or "",
                "Tienda": p.store,
                "URL": p.url,
                "Precio Actual": p.current_price or 0,
                "Moneda": p.currency,
                "Ultima Actualizacion": p.last_updated.strftime("%Y-%m-%d %H:%M") if p.last_updated else "",
            })

        df = pd.DataFrame(data)
        df.to_csv(filepath, index=False, encoding="utf-8-sig")
        logger.info("Exportación CSV completada: %s", filepath)
